from openpyxl import load_workbook

wb = load_workbook('_SCCO.xlsx')

# grab the active worksheet
ws = wb.active
start_row = 2
# Data can be assigned directly to cells
data = [['CCO/37/207/5/29/233/2018', 100, 'Butere', '30-Oct-18', 'Parental Child abduction',
         'Female', 2, 'Referred to other Government agencies', 'Yes', '01-Jan-12']]
for x in range(0, len(data)):
    for i in range(0, 9):
        e = ws.cell(row=start_row, column=i + 1)
        e.value = data[x][i]
    # Save the file

start_row += 1
wb.save("sample.xlsx")
